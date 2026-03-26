"""
excel_dashboard_multi_agent.py

Multi-agent Excel dashboard builder using enki_py.

Five specialist agents collaborate via MultiAgentRuntime:
  1. Requirements Agent  – parses natural-language dashboard requirements
  2. Planner Agent       – creates a structured dashboard layout plan
  3. Review Agent        – validates the plan for completeness
  4. Excel Agent         – analyses an input workbook and builds the dashboard
  5. Orchestrator        – top-level coordinator that delegates to the others

The workbook layer uses openpyxl to read source data and write a dashboard-
oriented workbook with KPI rows, chart-ready aggregated sections, and
planning notes.

Usage:
    python excel_dashboard_multi_agent.py \
        --requirements "Create a sales dashboard with revenue KPIs, region comparison, and top products" \
        --input sales_data.xlsx \
        --output sales_dashboard.xlsx
"""

from __future__ import annotations

import argparse
import json
import os
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from enki_py import Agent, MultiAgentMember, MultiAgentRuntime

model_name = "anthropic::claude-sonnet-4-6"
MAX_PREVIEW_ROWS = 3


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class SheetSummary:
    name: str
    row_count: int
    columns: list[str]
    sample_rows: list[dict[str, Any]]
    numeric_columns: list[str]


@dataclass
class KpiRow:
    label: str
    value: float | str


@dataclass
class AggregateSection:
    title: str
    headers: list[str]
    rows: list[list[str | float]]


@dataclass
class DashboardData:
    kpis: list[KpiRow]
    sections: list[AggregateSection]
    planning_notes: list[str]


@dataclass
class AgentContext:
    """Shared mutable state passed between agents via tool closures."""

    requirements: str = ""
    input_path: str = ""
    output_path: str = "dashboard_output.xlsx"
    sheet_summaries: list[SheetSummary] = field(default_factory=list)
    parsed_requirements: str = ""
    dashboard_plan: str = ""
    review_feedback: str = ""
    dashboard_data: DashboardData | None = None


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------


def summarise_workbook(file_path: str) -> list[SheetSummary]:
    """Read an Excel file and return per-sheet summaries."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    summaries: list[SheetSummary] = []

    for name in wb.sheetnames:
        ws = wb[name]
        rows_data: list[dict[str, Any]] = []
        columns: list[str] = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                columns = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                continue
            rows_data.append(dict(zip(columns, row)))

        # Detect numeric columns from first 20 data rows
        numeric_columns = [
            col
            for col in columns
            if any(isinstance(r.get(col), (int, float)) for r in rows_data[:20])
        ]

        summaries.append(
            SheetSummary(
                name=name,
                row_count=len(rows_data),
                columns=columns,
                sample_rows=rows_data[:5],
                numeric_columns=numeric_columns,
            )
        )

    wb.close()
    return summaries


def parse_requirement_text(text: str) -> dict[str, Any]:
    """Extract a compact structured requirement set from free text."""
    lower = text.lower()

    kpis: list[str] = []
    charts: list[str] = []
    dimensions: list[str] = []

    kpi_patterns = ["revenue", "sales", "profit", "growth", "margin", "count", "total", "average"]
    for p in kpi_patterns:
        if p in lower:
            kpis.append(p)

    chart_patterns = ["comparison", "trend", "distribution", "top", "breakdown", "pie", "bar", "line"]
    for p in chart_patterns:
        if p in lower:
            charts.append(p)

    dim_patterns = ["region", "product", "category", "date", "month", "quarter", "year", "customer"]
    for p in dim_patterns:
        if p in lower:
            dimensions.append(p)

    if not kpis:
        kpis = ["total", "average"]
    if not charts:
        charts = ["bar", "comparison"]

    return {
        "kpis": kpis,
        "chartTypes": charts,
        "dimensions": dimensions,
        "rawRequirements": text,
    }


def build_dashboard_workbook(
        source_path: str | None,
        dashboard: DashboardData,
        output_path: str,
) -> None:
    """Write a dashboard-oriented workbook to disk."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Dashboard"

    current_row = 1

    # --- KPI section -------------------------------------------------------
    ws.cell(row=current_row, column=1, value="KEY PERFORMANCE INDICATORS")
    ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True, size=14)
    current_row += 1

    ws.cell(row=current_row, column=1, value="Metric")
    ws.cell(row=current_row, column=2, value="Value")
    for col in (1, 2):
        ws.cell(row=current_row, column=col).font = ws.cell(row=current_row, column=col).font.copy(bold=True)
    current_row += 1

    for kpi in dashboard.kpis:
        ws.cell(row=current_row, column=1, value=kpi.label)
        ws.cell(row=current_row, column=2, value=kpi.value)
        current_row += 1

    current_row += 1  # spacer

    # --- Aggregated sections (chart-ready) ---------------------------------
    for section in dashboard.sections:
        ws.cell(row=current_row, column=1, value=section.title)
        ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True, size=12)
        current_row += 1

        for col_idx, header in enumerate(section.headers, start=1):
            ws.cell(row=current_row, column=col_idx, value=header)
            ws.cell(row=current_row, column=col_idx).font = ws.cell(row=current_row, column=col_idx).font.copy(
                bold=True)
        current_row += 1

        for data_row in section.rows:
            for col_idx, val in enumerate(data_row, start=1):
                ws.cell(row=current_row, column=col_idx, value=val)
            current_row += 1

        current_row += 1  # spacer

    # --- Planning notes ----------------------------------------------------
    if dashboard.planning_notes:
        ws.cell(row=current_row, column=1, value="PLANNING NOTES")
        ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True, size=12)
        current_row += 1
        for note in dashboard.planning_notes:
            ws.cell(row=current_row, column=1, value=note)
            current_row += 1

    # Auto-size columns
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_val in row_cells:
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

    # --- Copy source data sheets -------------------------------------------
    if source_path and os.path.exists(source_path):
        source_wb = load_workbook(source_path, read_only=True, data_only=True)
        for src_name in source_wb.sheetnames:
            src_ws = source_wb[src_name]
            dest_ws = wb.create_sheet(title=f"Source_{src_name}"[:31])
            for row in src_ws.iter_rows(values_only=True):
                dest_ws.append(list(row))
        source_wb.close()

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Agent factory
# ---------------------------------------------------------------------------


def create_agents(
        ctx: AgentContext, model: str
) -> tuple[Agent, Agent, Agent, Agent]:
    """Create the four specialist agents with tools wired to *ctx*."""

    # ------- Requirements Agent --------------------------------------------
    requirements_agent = Agent(
        model,
        name="RequirementsAgent",
        instructions=(
            "You are a requirements analyst for Excel dashboard projects. "
            "When asked, use the parse_requirements tool to extract structured "
            "requirements from natural language. Return only the parsed output."
        ),
    )

    @requirements_agent.tool_plain
    def parse_requirements(text: str) -> str:
        """Parse natural-language dashboard requirements into structured KPIs, chart types, and dimensions."""
        result = parse_requirement_text(text)
        ctx.parsed_requirements = json.dumps(result, separators=(",", ":"))
        return ctx.parsed_requirements

    # ------- Planner Agent -------------------------------------------------
    planner_agent = Agent(
        model,
        name="PlannerAgent",
        instructions=(
            "You are a dashboard layout planner. "
            "Use the create_dashboard_plan tool to generate a structured plan. "
            "Combine the parsed requirements with available data columns. "
            "Return only the plan JSON."
        ),
    )

    @planner_agent.tool_plain
    def create_dashboard_plan(requirements: str, available_columns: str = "[]") -> str:
        """Generate a dashboard plan from parsed requirements and column names."""
        try:
            reqs = json.loads(requirements)
        except (json.JSONDecodeError, TypeError):
            reqs = parse_requirement_text(requirements or ctx.requirements)

        numeric_cols = [c for s in ctx.sheet_summaries for c in s.numeric_columns]

        kpi_defs = []
        for kpi in reqs.get("kpis", ["total"]):
            col = next((c for c in numeric_cols if kpi in c.lower()), None) or (
                numeric_cols[0] if numeric_cols else "value")
            kpi_defs.append({"label": kpi.capitalize(), "sourceColumn": col, "aggregation": "sum"})

        chart_sections = []
        for chart_type in reqs.get("chartTypes", ["bar"]):
            group_col = None
            for d in reqs.get("dimensions", []):
                for s in ctx.sheet_summaries:
                    if any(d in c.lower() for c in s.columns):
                        group_col = next(c for c in s.columns if d in c.lower())
                        break
                if group_col:
                    break
            if not group_col:
                group_col = ctx.sheet_summaries[0].columns[0] if ctx.sheet_summaries and ctx.sheet_summaries[
                    0].columns else "category"

            value_col = numeric_cols[0] if numeric_cols else "value"
            chart_sections.append({
                "chartType": chart_type,
                "title": f"{chart_type.capitalize()} Chart",
                "groupByColumn": group_col,
                "valueColumn": value_col,
                "aggregation": "sum",
            })

        plan = {
            "kpis": kpi_defs,
            "charts": chart_sections,
            "sourceSheets": [s.name for s in ctx.sheet_summaries],
        }
        ctx.dashboard_plan = json.dumps(plan, separators=(",", ":"))
        return ctx.dashboard_plan

    # ------- Review Agent --------------------------------------------------
    review_agent = Agent(
        model,
        name="ReviewAgent",
        instructions=(
            "You are a dashboard plan reviewer. "
            "Use the review_plan tool to validate the dashboard plan. "
            "Report any issues or confirm readiness."
        ),
    )

    @review_agent.tool_plain
    def review_plan(plan: str) -> str:
        """Validate a dashboard plan for completeness and data alignment."""
        issues: list[str] = []
        available_cols = {
            c.lower() for s in ctx.sheet_summaries for c in s.columns
        }

        try:
            parsed_plan = json.loads(plan)
        except (json.JSONDecodeError, TypeError):
            return json.dumps({"approved": False, "issues": ["Plan JSON is malformed."]})

        for kpi in parsed_plan.get("kpis", []):
            src = kpi.get("sourceColumn", "")
            if src and src.lower() not in available_cols:
                issues.append(f'KPI source column "{src}" not found in data.')

        for chart in parsed_plan.get("charts", []):
            grp = chart.get("groupByColumn", "")
            val = chart.get("valueColumn", "")
            if grp and grp.lower() not in available_cols:
                issues.append(f'Chart group column "{grp}" not found in data.')
            if val and val.lower() not in available_cols:
                issues.append(f'Chart value column "{val}" not found in data.')

        approved = len(issues) == 0
        result = {
            "approved": approved,
            "issues": issues,
            "notes": "Plan looks good – proceed with build." if approved else "Fix the issues above before building.",
        }
        ctx.review_feedback = json.dumps(result, separators=(",", ":"))
        return ctx.review_feedback

    # ------- Excel Agent (analyser + builder) ------------------------------
    excel_agent = Agent(
        model,
        name="ExcelAgent",
        instructions=(
            "You are an Excel dashboard builder. "
            "Use analyse_workbook to inspect the source data. "
            "Use build_dashboard to create the output workbook. "
            "Return the build summary."
        ),
    )

    @excel_agent.tool_plain
    def analyse_workbook(file_path: str = "") -> str:
        """Analyse an Excel workbook and return sheet summaries with column types."""
        p = file_path or ctx.input_path
        if not p or not os.path.exists(p):
            return json.dumps({"error": f"File not found: {p}", "usingDemoData": True})

        ctx.sheet_summaries = summarise_workbook(p)
        return json.dumps(
            [
                {
                    "name": s.name,
                    "rowCount": s.row_count,
                    "columns": s.columns,
                    "numericColumns": s.numeric_columns,
                }
                for s in ctx.sheet_summaries
            ],
            default=str,
            separators=(",", ":"),
        )

    @excel_agent.tool_plain
    def build_dashboard(plan: str = "", output_path: str = "") -> str:
        """Build the dashboard workbook from the plan and write it to disk."""
        out_path = output_path or ctx.output_path

        try:
            parsed_plan = json.loads(plan or ctx.dashboard_plan)
        except (json.JSONDecodeError, TypeError):
            parsed_plan = {"kpis": [], "charts": []}

        # Collect all source rows
        all_rows: list[dict[str, Any]] = []
        for s in ctx.sheet_summaries:
            if ctx.input_path and os.path.exists(ctx.input_path):
                wb = load_workbook(ctx.input_path, read_only=True, data_only=True)
                ws = wb[s.name]
                columns: list[str] = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        columns = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                        continue
                    all_rows.append(dict(zip(columns, row)))
                wb.close()
            else:
                all_rows.extend(s.sample_rows)

        # Compute KPIs
        kpis: list[KpiRow] = []
        for k in parsed_plan.get("kpis", []):
            values = []
            for r in all_rows:
                try:
                    values.append(float(r.get(k["sourceColumn"], 0) or 0))
                except (ValueError, TypeError):
                    pass

            agg = k.get("aggregation", "sum")
            if agg == "average":
                value = round(sum(values) / len(values), 2) if values else 0
            elif agg == "count":
                value = len(values)
            elif agg == "min":
                value = min(values) if values else 0
            elif agg == "max":
                value = max(values) if values else 0
            else:  # sum
                value = round(sum(values), 2)
            kpis.append(KpiRow(label=k["label"], value=value))

        # Chart-ready aggregate sections
        sections: list[AggregateSection] = []
        for chart in parsed_plan.get("charts", []):
            grouped: dict[str, list[float]] = defaultdict(list)
            for row in all_rows:
                group_key = str(row.get(chart["groupByColumn"], "Unknown"))
                try:
                    grouped[group_key].append(float(row.get(chart["valueColumn"], 0) or 0))
                except (ValueError, TypeError):
                    pass

            headers = [chart["groupByColumn"], f"{chart['valueColumn']} ({chart.get('aggregation', 'sum')})"]
            data_rows: list[list[str | float]] = []
            for group, vals in grouped.items():
                if chart.get("aggregation") == "average":
                    agg_val = round(sum(vals) / len(vals), 2) if vals else 0
                else:
                    agg_val = round(sum(vals), 2)
                data_rows.append([group, agg_val])

            data_rows.sort(key=lambda r: float(r[1]), reverse=True)
            sections.append(AggregateSection(title=chart["title"], headers=headers, rows=data_rows))

        planning_notes = [
            f"Generated from: {ctx.input_path or 'demo data'}",
            f"Requirements: {ctx.requirements}",
            f"Plan review: {ctx.review_feedback or 'N/A'}",
        ]

        dashboard = DashboardData(kpis=kpis, sections=sections, planning_notes=planning_notes)
        ctx.dashboard_data = dashboard

        build_dashboard_workbook(ctx.input_path or None, dashboard, out_path)

        return json.dumps({
            "success": True,
            "outputPath": out_path,
            "kpiCount": len(kpis),
            "sectionCount": len(sections),
            "message": f"Dashboard written to {out_path}",
        })

    return requirements_agent, planner_agent, review_agent, excel_agent


# ---------------------------------------------------------------------------
# Multi-agent orchestrator members
# ---------------------------------------------------------------------------

DEMO_SHEET_SUMMARIES = [
    SheetSummary(
        name="SalesData",
        row_count=100,
        columns=["Region", "Product", "Revenue", "Units", "Date"],
        sample_rows=[
            {"Region": "North", "Product": "Widget A", "Revenue": 15000, "Units": 150, "Date": "2025-01"},
            {"Region": "South", "Product": "Widget B", "Revenue": 22000, "Units": 200, "Date": "2025-01"},
            {"Region": "East", "Product": "Widget A", "Revenue": 18000, "Units": 180, "Date": "2025-02"},
            {"Region": "West", "Product": "Widget C", "Revenue": 12000, "Units": 100, "Date": "2025-02"},
            {"Region": "North", "Product": "Widget B", "Revenue": 25000, "Units": 250, "Date": "2025-03"},
        ],
        numeric_columns=["Revenue", "Units"],
    )
]


def create_orchestrator_members(
        model: str,
        requirements_agent: Agent,
        planner_agent: Agent,
        review_agent: Agent,
        excel_agent: Agent,
) -> list[MultiAgentMember]:
    """Build MultiAgentMember list for the runtime registry."""
    orchestrator = Agent(
        model,
        name="Orchestrator",
        instructions=(
            "You are the orchestrator for an Excel dashboard builder pipeline. "
            "Use discover_agents to find specialist agents. "
            "Delegate tasks in this order: "
            "1. Delegate to ExcelAgent to analyse the workbook. "
            "2. Delegate to RequirementsAgent to parse the requirements. "
            "3. Delegate to PlannerAgent to create a dashboard plan. "
            "4. Delegate to ReviewAgent to validate the plan. "
            "5. Delegate to ExcelAgent to build the dashboard. "
            "Return a summary of the pipeline results."
        ),
    )

    return [
        MultiAgentMember(
            agent_id="orchestrator",
            agent=orchestrator,
            capabilities=["orchestration", "planning"],
            description="Routes work across specialist agents.",
        ),
        MultiAgentMember(
            agent_id="requirements-agent",
            agent=requirements_agent,
            capabilities=["requirements", "analysis"],
            description="Parses natural-language dashboard requirements.",
        ),
        MultiAgentMember(
            agent_id="planner-agent",
            agent=planner_agent,
            capabilities=["planning", "layout"],
            description="Creates structured dashboard layout plans.",
        ),
        MultiAgentMember(
            agent_id="review-agent",
            agent=review_agent,
            capabilities=["review", "validation"],
            description="Validates dashboard plans for completeness.",
        ),
        MultiAgentMember(
            agent_id="excel-agent",
            agent=excel_agent,
            capabilities=["excel", "data-analysis", "dashboard-building"],
            description="Analyses workbooks and builds dashboards.",
        ),
    ]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel Dashboard Multi-Agent Builder")
    parser.add_argument(
        "--requirements",
        default="Create a sales dashboard with revenue KPIs, region comparison, and top products",
        help="Natural-language dashboard requirements",
    )
    parser.add_argument("--input", dest="input_path", default="", help="Path to source .xlsx file")
    parser.add_argument("--output", dest="output_path", default="dashboard_output.xlsx", help="Output .xlsx path")
    parser.add_argument("--model", default=None,
                        help="Model identifier (default: ENKI_MODEL env or ollama::qwen3.5:latest)")
    args = parser.parse_args()

    model = args.model or os.environ.get("ENKI_MODEL", model_name)

    print("╔════════════════════════════════════════════════════════════╗")
    print("║         Excel Dashboard Multi-Agent Builder (Python)      ║")
    print("╚════════════════════════════════════════════════════════════╝")
    print()
    print(f"  Model:        {model}")
    print(f"  Input:        {args.input_path or '(no input file – using demo data)'}")
    print(f"  Output:       {args.output_path}")
    print(f"  Requirements: {args.requirements}")
    print()

    ctx = AgentContext(
        requirements=args.requirements,
        input_path=str(Path(args.input_path).resolve()) if args.input_path else "",
        output_path=args.output_path,
    )

    requirements_agent, planner_agent, review_agent, excel_agent = create_agents(ctx, model)
    session_id = "excel-dashboard-session"

    # --- Stage 1: Analyse input workbook -----------------------------------
    print("─── Stage 1: Analyse input workbook ────────────────────────")
    if ctx.input_path and os.path.exists(ctx.input_path):
        result = excel_agent.run_sync(
            f'Analyse the workbook at "{ctx.input_path}" using the analyse_workbook tool.',
            session_id=session_id,
        )
        print(result.output)
    else:
        print("  No input file provided – will use demo data.")
        ctx.sheet_summaries = list(DEMO_SHEET_SUMMARIES)
    print()

    # --- Stage 2: Parse requirements ---------------------------------------
    print("─── Stage 2: Parse requirements ────────────────────────────")
    result = requirements_agent.run_sync(
        f'Use the parse_requirements tool on this dashboard request: "{args.requirements}"',
        session_id=session_id,
    )
    print(result.output)
    print()

    # --- Stage 3: Create dashboard plan ------------------------------------
    print("─── Stage 3: Create dashboard plan ─────────────────────────")
    all_columns = [c for s in ctx.sheet_summaries for c in s.columns]
    result = planner_agent.run_sync(
        "Use the create_dashboard_plan tool to create a dashboard plan. "
        f"Pass requirements={json.dumps(ctx.requirements)} and "
        f"available_columns={json.dumps(json.dumps(all_columns, separators=(',', ':')))}.",
        session_id=session_id,
    )
    print(result.output)
    print()

    # --- Stage 4: Review plan ----------------------------------------------
    print("─── Stage 4: Review plan ───────────────────────────────────")
    result = review_agent.run_sync(
        "Use the review_plan tool to validate the current dashboard plan stored in context. "
        f"Pass plan={json.dumps(ctx.dashboard_plan)}.",
        session_id=session_id,
    )
    print(result.output)
    print()

    # --- Stage 5: Build dashboard workbook ---------------------------------
    print("─── Stage 5: Build dashboard workbook ──────────────────────")
    result = excel_agent.run_sync(
        "Use the build_dashboard tool with the current plan stored in context. "
        f"Pass output_path={json.dumps(ctx.output_path)}.",
        session_id=session_id,
    )
    print(result.output)
    print()

    # --- Multi-agent runtime (shows registry + discovery) ------------------
    print("─── Agent Registry (MultiAgentRuntime) ─────────────────────")
    members = create_orchestrator_members(model, requirements_agent, planner_agent, review_agent, excel_agent)
    runtime = MultiAgentRuntime(members)

    cards = runtime.registry()
    print("  Registered agents:")
    for card in cards:
        print(f"    • {card.agent_id} ({card.name}) – capabilities: {card.capabilities}")

    excel_cards = runtime.discover(capability="excel")
    print(f"\n  Excel-capable agents: {', '.join(c.name for c in excel_cards)}")
    print()

    print("═════════════════════════════════════════════════════════════")
    print(f"  ✓ Dashboard written to: {ctx.output_path}")
    print("═════════════════════════════════════════════════════════════")
    print()

    # --- Interactive Q&A loop ---------------------------------------------
    _run_interactive_loop(ctx, runtime, session_id)


def _run_interactive_loop(
    ctx: AgentContext,
    runtime: MultiAgentRuntime,
    session_id: str,
) -> None:
    """Interactive REPL where the user can ask follow-up questions.

    Questions are routed through the orchestrator, which can delegate to
    any specialist agent (requirements, planner, review, excel) as needed.
    """

    # Build a context summary the orchestrator can reference
    data_summary = _build_data_summary(ctx)

    print("┌────────────────────────────────────────────────────────────┐")
    print("│  Interactive Mode – ask questions about your dashboard     │")
    print("│                                                            │")
    print("│  Examples:                                                 │")
    print('│    • "What are the top 3 regions by revenue?"              │')
    print('│    • "Explain the KPIs in the dashboard"                   │')
    print('│    • "Rebuild the dashboard with average instead of sum"   │')
    print('│    • "What columns are available in the source data?"      │')
    print("│                                                            │")
    print("│  Type 'quit' or 'exit' to leave.                           │")
    print("└────────────────────────────────────────────────────────────┘")
    print()

    question_num = 0
    while True:
        try:
            user_input = input("❯ ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n  Goodbye!")
            break

        if not user_input:
            continue
        if user_input.lower() in ("quit", "exit", "q"):
            print("  Goodbye!")
            break

        question_num += 1
        q_session = f"{session_id}-q{question_num}"

        # Compose the full prompt with data context
        prompt = (
            f"The user has built a dashboard and is now asking a follow-up question.\n\n"
            f"--- Data Context ---\n{data_summary}\n\n"
            f"--- User Question ---\n{user_input}\n\n"
            f"Answer the question using the available data context. "
            f"If the user asks to rebuild or modify the dashboard, use discover_agents "
            f"and delegate to the appropriate specialist agent. "
            f"Keep your answer concise and helpful."
        )

        print()
        try:
            result = runtime.process_sync(
                "orchestrator",
                prompt,
                session_id=q_session,
            )
            print(f"  {result.output}")
        except Exception as exc:
            print(f"  ⚠ Error: {exc}")
        print()


def _build_data_summary(ctx: AgentContext) -> str:
    """Build a concise text summary of the current dashboard state for the LLM."""
    parts: list[str] = []

    parts.append(f"Requirements: {ctx.requirements}")
    parts.append(f"Input file: {ctx.input_path or 'demo data'}")
    parts.append(f"Output file: {ctx.output_path}")

    # Sheet info
    if ctx.sheet_summaries:
        sheets_info = []
        for s in ctx.sheet_summaries:
            preview_count = min(len(s.sample_rows), MAX_PREVIEW_ROWS)
            sheets_info.append(
                f"  Sheet '{s.name}': {s.row_count} rows, "
                f"columns={s.columns}, numeric={s.numeric_columns}, "
                f"sample_rows={preview_count}"
            )
        parts.append("Source sheets:\n" + "\n".join(sheets_info))

    # KPIs
    if ctx.dashboard_data and ctx.dashboard_data.kpis:
        kpi_lines = [f"  {k.label}: {k.value}" for k in ctx.dashboard_data.kpis]
        parts.append("Dashboard KPIs:\n" + "\n".join(kpi_lines))

    # Sections
    if ctx.dashboard_data and ctx.dashboard_data.sections:
        for section in ctx.dashboard_data.sections:
            rows_preview = section.rows[:5]
            section_lines = [f"  {section.headers}"]
            for row in rows_preview:
                section_lines.append(f"  {row}")
            if len(section.rows) > 5:
                section_lines.append(f"  ... ({len(section.rows) - 5} more rows)")
            parts.append(f"Section '{section.title}':\n" + "\n".join(section_lines))

    # Plan + review
    if ctx.dashboard_plan:
        try:
            plan = json.loads(ctx.dashboard_plan)
            parts.append(
                "Dashboard plan summary: "
                f"{len(plan.get('kpis', []))} KPIs, {len(plan.get('charts', []))} charts, "
                f"source sheets={plan.get('sourceSheets', [])}"
            )
        except (json.JSONDecodeError, TypeError):
            parts.append("Dashboard plan summary: available")
    if ctx.review_feedback:
        try:
            review = json.loads(ctx.review_feedback)
            parts.append(
                "Review feedback: "
                f"approved={review.get('approved')}, issues={len(review.get('issues', []))}"
            )
        except (json.JSONDecodeError, TypeError):
            parts.append("Review feedback: available")

    return "\n\n".join(parts)


if __name__ == "__main__":
    main()
